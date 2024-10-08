from openpyxl import load_workbook
import pytest
from base.get_yaml_data import HandleYaml
import os

tg_root = os.path.dirname(os.path.dirname(__file__))  # 获取当前文件所在目录的上级目录
excle_yaml_path = os.path.join(tg_root, r'test_datas\excle_config.yaml')
excle_path = os.path.join(tg_root, r'test_datas\test_data1.xlsx')
print(excle_path)
class Handle_excle:
    def __init__(self):
        # global self.sheet1
        try:
            workbook = load_workbook(excle_path)  # 打开excle文件
        except Exception:
            print("找不到文件")
        handle = HandleYaml()  # 读取yaml文件中的excl表名
        list_sheet = handle.load_yaml(excle_yaml_path)['sheet_name']
        sheet_name: str = ''.join(list_sheet)  # 转化为字符串
        self.sheet1 = workbook[sheet_name]  # 通过sheet名查找

    def get_cell_value(self, xrow: int, xcol: int): 
        """

        :type xrow: 第几行
        :type xcol: 第几列
        """
        cell_value = self.sheet1.cell(xrow, xcol).value  # 取第xx行第xx列
        return cell_value

    def get_nrow(self):
        nrow = self.sheet1.max_row
        return nrow   #获取表的行数

    def get_ncol(self):
        ncol = self.sheet1.max_column
        return ncol   #获取表的列数

    def row_datas(self, row):
        # 获取excle中某一行数据
        for row_cell in self.sheet1.iter_rows(min_row=row, max_row=row, values_only=True):  # 遍历某一行
            return row_cell
    
    def all_datas(self):
        for i in self.sheet1.iter_rows(min_row=2, min_col=2, values_only=True):
            yield i

hand_excl = Handle_excle()

data1 = list(hand_excl.all_datas())
# print(data1)
